#!/usr/bin/env python3
"""T5 Gate 2.4 — regenerate the DB contract snapshot workbook from the LIVE DB.

Spec: docs/T5_guardrails_and_checks.md 2.4. The snapshot is the only artefact
that survives project knowledge (binaries are corrupted there), so it WILL be
read - and it must date itself. Requirements met here:
  - an as-of timestamp on EVERY sheet header (not just the README);
  - a "point-in-time; re-derive before acting" banner on every QA-derived sheet;
  - run at the end of every task Gate C to keep the project-knowledge copy current.

Everything is read straight from the live DB, so the workbook cannot drift from
it the way a hand-maintained copy did. Run:
  python scripts/utils/build_db_contract_snapshot.py            # writes docs/..._<date>.xlsx
  python scripts/utils/build_db_contract_snapshot.py OUT.xlsx   # explicit path
"""
from __future__ import annotations

import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
DB = ROOT / "Output" / "database" / "Gayini_Results.sqlite"

AS_OF = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
DATESTAMP = datetime.now(timezone.utc).strftime("%Y%m%d")

QA_BANNER = ("POINT-IN-TIME — re-derive before acting. QA/release verdicts may be "
             "stale (several date from 2026-07-01 and are contradicted by current "
             "data). Read via v_qa_freshness once it exists; never quote a QA row as current.")

BOLD = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="DDEBF7")
QA_FILL = PatternFill("solid", fgColor="FCE4D6")


def q(con, sql, params=()):
    cur = con.execute(sql, params)
    return [d[0] for d in cur.description], cur.fetchall()


def add_sheet(wb, name, header, rows, qa=False):
    ws = wb.create_sheet(name[:31])
    ws.append([f"as-of {AS_OF}  |  source: {DB.name}  |  sheet: {name}"])
    ws["A1"].font = BOLD
    if qa:
        ws.append([QA_BANNER])
        ws["A2"].font = Font(bold=True, color="C00000")
        ws["A2"].fill = QA_FILL
    ws.append([])
    hrow = ws.max_row + 1
    ws.append(list(header))
    for c in ws[hrow]:
        c.font = BOLD
        c.fill = HDR_FILL
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    return ws


def main(out_path: str | None) -> None:
    if not DB.is_file():
        raise SystemExit(f"ABORT: {DB} not found.")
    out = Path(out_path) if out_path else ROOT / "docs" / f"Gayini_Results_DB_contract_snapshot_{DATESTAMP}.xlsx"
    con = sqlite3.connect(f"file:{DB.as_posix()}?mode=ro", uri=True)
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # README
        ws = wb.create_sheet("00_README")
        for line in [
            f"GAYINI RESULTS DATABASE — CONTRACT SNAPSHOT (machine-generated)",
            f"as-of {AS_OF}",
            f"source: {DB}  ({DB.stat().st_size:,} bytes)",
            "",
            "Regenerated from the LIVE DB by scripts/utils/build_db_contract_snapshot.py.",
            "Authoritative for object existence, schema and row counts. NOT for QA verdicts:",
            "the QA / release sheets are point-in-time and flagged; re-derive from data.",
            "Every sheet header carries its own as-of timestamp.",
        ]:
            ws.append([line])
        ws["A1"].font = BOLD

        # Objects: tables + views with row counts
        names = con.execute(
            "SELECT name, type FROM sqlite_master WHERE type IN ('table','view') "
            "AND name NOT LIKE 'sqlite_%' ORDER BY type, name").fetchall()
        obj_rows = []
        for nm, typ in names:
            try:
                n = con.execute(f"SELECT COUNT(*) FROM '{nm}'").fetchone()[0]
            except Exception as e:
                n = f"ERR: {e}"
            obj_rows.append([typ, nm, n])
        add_sheet(wb, "01_Objects", ["type", "name", "row_count"], obj_rows)

        # Schema: every column
        sch = []
        for nm, typ in names:
            for cid, cname, ctype, notnull, dflt, pk in con.execute(f"PRAGMA table_info('{nm}')"):
                sch.append([typ, nm, cid, cname, ctype, notnull, pk])
        add_sheet(wb, "02_Schema", ["obj_type", "object", "cid", "column", "dtype", "notnull", "pk"], sch)

        # Full-table dumps for the small, verification-critical registries + T1 objects
        for sheet, table, order in [
            ("03_Spatial_layers", "spatial_layer_asset", "spatial_layer_asset_id"),
            ("04_Census_asset", "census_asset", "census_asset_id"),
            ("05_Dim_management_zone", "dim_management_zone", "zone_fid"),
            ("06_T1_zone_identity", "t1_zone_identity_check", "check_id"),
            ("07_Dim_spatial_unit", "dim_spatial_unit", "unit_id"),
            ("08_Census_stratum", "census_stratum", "rowid"),
            ("09_Metrics", "dim_metric", "rowid"),
        ]:
            try:
                header, rows = q(con, f"SELECT * FROM {table} ORDER BY {order}")
                add_sheet(wb, sheet, header, rows)
            except Exception as e:
                add_sheet(wb, sheet, ["error"], [[f"{table}: {e}"]])

        # Raster assets — key columns
        header, rows = q(con,
            "SELECT raster_asset_id, product, crs_epsg, resolution_x, xmin, ymin, xmax, ymax, "
            "legend_status, path_exists FROM raster_asset ORDER BY product, raster_asset_id")
        add_sheet(wb, "10_Raster_assets", header, rows)

        # Figure assets — T1 rows + a run_id summary
        header, rows = q(con,
            "SELECT figure_asset_id, support_level, figure_level, run_id, checksum_sha256 "
            "FROM figure_asset WHERE run_id LIKE 'T1_%' ORDER BY figure_asset_id")
        add_sheet(wb, "11_Figure_assets_T1", header, rows)

        # Release checks + QA issues — flagged as point-in-time
        for sheet, view in [("12_Release_checks", "v_database_release_checks"),
                            ("13_QA_issues", "v_current_qa_issues")]:
            try:
                header, rows = q(con, f"SELECT * FROM {view}")
                add_sheet(wb, sheet, header, rows, qa=True)
            except Exception as e:
                add_sheet(wb, sheet, ["error"], [[str(e)]], qa=True)

        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"[snapshot] wrote {out}")
        print(f"[snapshot] as-of {AS_OF}; {len(names)} objects; "
              f"{con.execute('SELECT COUNT(*) FROM dim_management_zone').fetchone()[0]} zones; "
              f"spatial_layer_asset={con.execute('SELECT COUNT(*) FROM spatial_layer_asset').fetchone()[0]}; "
              f"figure_asset={con.execute('SELECT COUNT(*) FROM figure_asset').fetchone()[0]}")
    finally:
        con.close()


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
