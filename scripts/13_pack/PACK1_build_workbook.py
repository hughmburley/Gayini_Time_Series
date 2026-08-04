#!/usr/bin/env python
"""PACK-1 P4-4..P4-7 + Ruling N — 00_START_HERE.md, the workbook, the number check, registration.

EVERY number on How_we_know is QUERIED LIVE (P4-6 / AD-B). Nothing on that sheet is typed.
Contents and 00_START_HERE.md both generate from PACK1_item_list.csv - one source, so they
cannot disagree (P4-4).
"""
import sqlite3, csv, hashlib, datetime, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "Output" / "pack"
DB = ROOT / "Output" / "database" / "Gayini_Results.sqlite"
RUN = "PACK1_P4_20260803"

def probe(lbl):
    c = sqlite3.connect(f"file:{DB.as_posix()}?mode=ro", uri=True); c.execute("PRAGMA query_only=1")
    o = {t: c.cursor().execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
         for t in ("dim_headline_number","figure_asset","raster_asset","table_asset","report_asset")}
    c.close(); print(f"  PROBE {lbl}: " + " · ".join(f"{k}={v}" for k,v in o.items())); return o

def sha50(p):
    h=hashlib.sha256(); cap=50*1024*1024; n=0
    with open(p,"rb") as f:
        while n<cap:
            b=f.read(1<<20)
            if not b: break
            h.update(b); n+=len(b)
    return h.hexdigest()

pre = probe("before")
con = sqlite3.connect(f"file:{DB.as_posix()}?mode=ro", uri=True); con.execute("PRAGMA query_only=1")
c = con.cursor()
PINS = dict(c.execute("SELECT number_id, pinned_value FROM dim_headline_number "
                      "WHERE pinned_value IS NOT NULL"))
items = sorted(csv.DictReader(open(PACK/"PACK1_item_list.csv", encoding="utf-8")),
               key=lambda r: int(r["display_order"]))   # FIG1-T1
audit = list(csv.DictReader(open(ROOT/"Output/tables/RPTSCOPE_claim_audit.csv", encoding="utf-8")))

# ---------- P4-6: EVERY figure here is a LIVE QUERY. None is typed. -------------------------
LIVE = {}
LIVE["registered"] = c.execute("SELECT COUNT(*) FROM dim_headline_number").fetchone()[0]
LIVE["pinned"]     = c.execute("SELECT COUNT(*) FROM dim_headline_number "
                               "WHERE pinned_value IS NOT NULL").fetchone()[0]
st = list(csv.DictReader(open(ROOT/"Output/tables/RPTSCOPE_reproduction_status.csv", encoding="utf-8")))
LIVE["reproduce"]  = sum(1 for r in st if r["status"]=="REPRODUCES")
LIVE["drift"]      = sum(1 for r in st if r["status"]=="VALUE_DRIFT")
LIVE["nopath"]     = sum(1 for r in st if r["status"]=="NO_DERIVATION_PATH")
LIVE["coverage"]   = 100*LIVE["reproduce"]/LIVE["pinned"]
LIVE["figures"]    = c.execute("SELECT COUNT(*) FROM figure_asset").fetchone()[0]
LIVE["tables"]     = c.execute("SELECT COUNT(*) FROM table_asset").fetchone()[0]
# RT-2: three registries are written by this build, not two. report_asset moved 59->60 in the SAME
# transaction that moved table_asset 4->5 (the write I-44 was about) and was silently absent.
LIVE["reports"]    = c.execute("SELECT COUNT(*) FROM report_asset").fetchone()[0]
print(f"  LIVE: registered={LIVE['registered']} pinned={LIVE['pinned']} reproduce={LIVE['reproduce']} "
      f"drift={LIVE['drift']} coverage={LIVE['coverage']:.1f}%")

# ---------- claims: text is design-seat verbatim, numbers carry their number_id -------------
CLAIMS = [
 ("1","Removing grazing has not, by itself, produced a measurably different floor. Across 35 years "
      "the three conserved paddocks other than Bala 29ca sat on average 2.1 percentage points below "
      "the grazed median at their vegetation floor, and in individual years ranged from 7.0 points "
      "below it to 5.0 points above. The difference crosses zero.",
      "ref_grazed_gap_annual_ref3_excl29ca_mean"),
 ("2","Bala 29ca's floor has been converging on the grazed median since 1988 at +0.92 percentage "
      "points a year. Conservation management began in 2019. The trend predates it by three decades.",
      "t10_gap_annual_slope_C_29ca"),
 ("3","How often a paddock floods correlates with its floor at r = 0.71 across all 64 paddocks — "
      "about half the variation between paddocks.", "floor_flood_r_64pdk"),
 ("4","82% of Bala 29ca's improvement survives removing the effect of the water it actually received, "
      "and the improvement is located in its dry western parts, not the paddock as a whole.",
      "bala29ca_improvement_surviving_water_pct"),
 ("5","The four conserved paddocks are not a matched comparison set. Ranked by how often they flood, "
      "they sit 3rd, 6th, 31st and 61st of 64 — almost the entire wetness range of the property. They "
      "are not one condition, and no single number describes them.",
      "ref_paddock_flood_rank_bala26ca,ref_paddock_flood_rank_bala28ca,"
      "ref_paddock_flood_rank_bala27ca,ref_paddock_flood_rank_bala29ca"),
 ("6","Eight of 115 paddock-parts are improving faster than their water explains while sitting below "
      "their community's typical floor. Five of the eight survive dropping the two wettest years. "
      "Three recover at every cut from 0.50 to 1.50, and two of those three are in Bala 29ca.",
      "t13_parts_recovering_count,t13_recovering_survive_drop2wettest"),
 ("7","Twelve of the sixteen declining parts are in the Bala group.", "t13_parts_declining_count"),
]
Q2 = ("No, and the more useful answer is that this design cannot settle it. Across 35 years the three "
      "conserved paddocks other than Bala 29ca sat 2.1 points below the grazed median on average, "
      "ranging from 7.0 below to 5.0 above — the difference crosses zero. But the four conserved "
      "paddocks span flood ranks 3, 6, 31 and 61 of 64, so they are not one condition and cannot serve "
      "as a single reference. All four are also in the Bala block, so removing grazing is perfectly "
      "confounded with where the paddocks are. Anything measured against this set is measured against "
      "a moving target.")
ANSWERS = {
 "Q1":("Are the formerly-cropped paddocks becoming more like the conserved ones?",
       "NOT ANSWERABLE AS ASKED","Cropping history is not recorded anywhere. Five columns are reserved "
       "for it and are empty for all 64 paddocks, so every contrast in this pack is not-grazed versus "
       "grazed, never conserved versus formerly-cropped.","cropping_history_null_count"),
 # RT-1: the cell quotes all four ranks and listed one. All four now listed.
 "Q2":("Are the conserved paddocks a usable reference set?","NO — AND THE DESIGN CANNOT SETTLE IT",Q2,
       "ref_grazed_gap_annual_ref3_excl29ca_mean,ref_paddock_flood_rank_bala26ca,"
       "ref_paddock_flood_rank_bala28ca,ref_paddock_flood_rank_bala27ca,"
       "ref_paddock_flood_rank_bala29ca"),
 "Q3":("Is Bala 29ca recovering, and is it just getting wetter?","RECOVERING, AND NOT JUST WETTER",
       "Its poorest patches carry about 17 percentage points less cover than its dryness predicts — the "
       "second largest shortfall on the property — and 82% of its improvement survives removing the "
       "effect of the water it actually received.",
       # RT-1: the "about 17 percentage points" was written but unlisted
       "bala29ca_improvement_surviving_water_pct,t10_bala29ca_xsec_residual"),
 "Q4":("Does grazing intensity show up in the ground cover?","NO — AND THE ORDERING RUNS THE WRONG WAY",
       "Comparing three management types within similar country, the standard-grazing land sits at or "
       "above the rotationally grazed land in six of nine comparisons. Either intensity does not "
       "register in this measure, or the unzoned country is grazed less rather than more.",
       "three_arm_standard_at_or_above_count"),
 "Q5":("What does drive ground cover?","WATER, AND IT IS NOT CLOSE",
       "Across all 64 paddocks, how often a paddock floods correlates with its floor at r = 0.71 — "
       "about half the variation between paddocks. The pattern follows channels and low ground and "
       "crosses paddock fences without noticing them.","floor_flood_r_64pdk"),
 "Q6":("Which country is coming back, and which is going backwards?",
       "BOTH, AND IT IS GEOGRAPHIC RATHER THAN MANAGERIAL",
       "Eight of 115 parts are improving faster than their water explains; five survive dropping the "
       "two wettest years; three recover at every cut tested. Twelve of the sixteen declining parts "
       "are in the Bala group.",
       # RT-1: "five survive dropping the two wettest years" was written but unlisted
       "t13_parts_recovering_count,t13_recovering_survive_drop2wettest,t13_parts_declining_count"),
 "Q7":("Did management change the water regime?","UNTESTED, AND PROBABLY UNTESTABLE WITH THIS RECORD",
       "This is the question worth asking and we cannot answer it. There are four water years since "
       "management changed and they are unusually wet. This is an honest non-answer, not a gap in the "
       "work: no design on this record can separate a management effect from that.","(none — N/A by design)"),
}
CAUTIONS = [
 ("Support levels are never merged","A plot-support number and a pixel-support number are different "
  "measurements of the same idea and are never compared or combined."),
 ("Cover is not condition","The floor says how much ground cover there is and how green it is. It is "
  "not a condition score and implies no cause."),
 ("A paddock is not an ecological unit","Management zones were drawn for stock and water. Decompose "
  "by vegetation community before attributing anything to a fence line."),
 ("No p-values on the annual series","Thirty-five consecutive years are not independent observations."),
 ('"The floor" always means one thing here','In this pack "the floor" always means the total-cover '
  'floor (veg_p05_spatial), never the green-share floor. The two are different quantities.'),
]

# ---------- P4-4: 00_START_HERE.md, generated ------------------------------------------------
def q_for(i):
    for k,(qq,ans,why,nid) in ANSWERS.items():
        if i["item_id"] in ("M1","F1","T1") and k=="Q1": return qq
    return ""
lines = ["# Start here", "", "**Gayini reference-state assessment — pack for Adrian · 10 August 2026**", "",
         "This folder is a derived artefact. Every file in it is copied from a registered source and",
         "verified by checksum on both sides; nothing was edited in place. The item-to-filename mapping",
         "below is **generated from `PACK1_item_list.csv`**, which is also what the workbook's Contents",
         "sheet reads — so the two cannot disagree. Filenames were deliberately **not** renamed: the pack",
         "path and the registry path are identical, so no mapping can drift.", "",
         "Items are ordered **by the argument they make**, not by item code: what bounds the pack and",
         "the data behind it first, then the argument itself, then the supporting detail.", ""]
WHAT = {"M1":"where the paddocks are","M2":"where the monitoring sites are",
 "M3":"which country stays green longest","M4":"which parts are coming back and which are going backwards",
 "M4b":"how that classification moves when the cut is loosened or tightened",
 "M5":"cover and water, at paddock and at part grain","M5b":"which paddocks beat or miss their water",
 "F1":"how each conserved paddock's floor tracks the grazed range","F2":"the same, on mean cover",
 "F3":"whether conserved country is pulling away from grazed country",
 "F4":"how the conserved-grazed gap narrowed","F5":"cover against water across all 64 paddocks",
 "F6":"whether grazing intensity shows up in the floor","F7":"Bala 29ca, part by part",
 "T1":"the four conserved paddocks side by side","T2":"every part of the property, to look up",
 "T3":"what this analysis cannot tell you","T1_render":"T1 as a picture"}
SECTION_WHY = {
 "Read first": "T3 bounds everything the rest of the pack can mean; T1 and T2 are the data the maps draw.",
 "The argument": "Where the country is, that cover follows water, the gap and who drives it, who beats "
                 "their water, which parts, and how much the answer moves when the cut does.",
 "Supporting detail": "Grain, decomposition, per-paddock trajectories, the three-arm comparison, "
                      "coverage and persistence."}
cur = None
for i in items:
    if i["section"] != cur:
        cur = i["section"]
        lines += ["", f"### {cur}", "", SECTION_WHY[cur], "",
                  "| item | file | what it answers |", "|---|---|---|"]
    name = Path(i["file_path"]).name if i["file_path"] else "—"
    lines.append(f"| **{i['item_id']}** | `{name}` | {WHAT.get(i['item_id'],'')} |")
lines += ["", f"*{len([x for x in items if x['item_id']!='T1_render'])} items in "
          f"{len({x['file_path'] for x in items if x['file_path']})} files. "
          f"F7 is a panel of M4's figure and shares its file; T1 ships as data with T1_render as its "
          f"picture.*", "",
          "**Read `Gayini_what_we_dont_know.md` before quoting anything from this pack.**"]
(PACK/"00_START_HERE.md").write_text("\n".join(lines), encoding="utf-8")
print(f"  wrote 00_START_HERE.md ({len(items)} rows, generated)")

# ---------- P4-5/P4-6: the workbook ----------------------------------------------------------
wb = openpyxl.Workbook(); wb.remove(wb.active)
B = Font(bold=True); W = Alignment(wrap_text=True, vertical="top")
def sheet(name, rows, widths):
    ws = wb.create_sheet(name)
    for r in rows: ws.append(r)
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width = w
    for cell in ws[1]: cell.font = B
    for row in ws.iter_rows():
        for cl in row: cl.alignment = W
    return ws

sheet("Start_here",
      [["#","claim","number_id (provenance)"]] + [[n,t,nid] for n,t,nid in CLAIMS],
      [5,110,46])
sheet("By_question",
      [["#","the question","the answer","why — in plain terms","number_id (provenance)"]] +
      [[k,v[0],v[1],v[2],v[3]] for k,v in ANSWERS.items()],
      [6,44,44,96,46])
ws_contents = sheet("Contents",
      [["#","section","item","file","what it answers","sha256 (source)","registered_in"]] +
      [[i["display_order"], i["section"], i["item_id"],
        Path(i["file_path"]).name if i["file_path"] else "—",
        WHAT.get(i["item_id"],""), (i["sha256"] or "")[:16], i["registered_in"]] for i in items],
      [5,20,10,52,54,20,46])
sheet("Two_cautions", [["caution","what it means"]] + [[a,b] for a,b in CAUTIONS], [42,104])

cov = (f"Of the {LIVE['pinned']} numbers in this pack's registry that carry a pinned value, "
       f"{LIVE['reproduce']} ({LIVE['coverage']:.1f}%) can be recomputed from source by a second, "
       f"independent route, and {LIVE['drift']} have drifted. The {LIVE['nopath']} not covered are "
       f"numbers for which no second route has been written — not numbers that failed.")
FALSIF = ("A rule written before the numbers arrived fired against us. It said that if the mapped "
          "area of persistent cover proved highly sensitive across plausible thresholds with no "
          "natural break, then 'refugia' is a chosen cut on a continuum and must be reported as a "
          "gradient rather than an area. It did, so no headline refugia figure was set. That is "
          "stronger evidence than the reproduction test: reproduction shows the numbers are stable, "
          "this shows they were not fitted.")
sheet("How_we_know",
      [["what","live value","how it is obtained"],
       ["numbers registered with a definition", LIVE["registered"], "COUNT(*) dim_headline_number — queried at build time"],
       ["of those, carrying a pinned value", LIVE["pinned"], "COUNT(*) WHERE pinned_value IS NOT NULL — queried at build time"],
       ["independently re-derived, and drifted", f"{LIVE['reproduce']} re-derived · {LIVE['drift']} drifted",
        "RPTSCOPE_reproduction_status.csv — read at build time; the two ALWAYS travel together"],
       ["coverage and drift, in one sentence", cov, "generated at build time; never typed, never copied"],
       ["registered figures", LIVE["figures"], "COUNT(*) figure_asset — queried at build time"],
       ["registered tables", LIVE["tables"], "COUNT(*) table_asset — queried at build time"],
       ["registered reports", LIVE["reports"], "COUNT(*) report_asset — queried at build time. This "
        "build writes all three registries; reporting two of them left the third silently absent"],
       ["a rule that fired against us", FALSIF, "T3 Gate B1 pre-registration; R2 Ruling D derivation"],
       ["", "", ""],
       ["NOTE", "No number on this sheet is typed.",
        "Every value is produced by a live query at build time. Counts move as work lands; a typed "
        "copy would be stale within the hour, and has been four times."]],
      [40,64,86])

# P4-4 / FIG1-T1: the two generated documents must agree on item, filename AND order. Derived
# from the rendered artefacts, not asserted - a criterion stated as a literal is not a check (I-40).
md_rows = [tuple(x.strip(" *`") for x in ln.strip("|").split("|")[:2])
           for ln in (PACK/"00_START_HERE.md").read_text(encoding="utf-8").splitlines()
           if ln.startswith("| **")]
xl_rows = [(r[2], r[3]) for r in ws_contents.iter_rows(min_row=2, values_only=True)]
assert md_rows == xl_rows, ("STOP - 00_START_HERE.md and Contents disagree"
                            f"\n  md: {md_rows}\n  xl: {xl_rows}")
print(f"  P4-4: 00_START_HERE.md and Contents agree on item, filename and order "
      f"({len(md_rows)} rows each)")

out = PACK/"Gayini_Adrian_pack.xlsx"
wb.save(out); print(f"  wrote {out.name} (5 sheets)")

# ---------- P4-7: final number check ---------------------------------------------------------
# RT-1 CHANGES THE RULE. The first version resolved every number_id LISTED IN THE PROVENANCE
# COLUMN, and passed 38/38 while three cells quoted numbers whose ids were not listed. A check
# that verifies the citation list rather than the sentence cannot see an uncited number - it is
# the I-40 shape again (the criterion and the thing checked wrong in the same direction).
# The rule is now: EXTRACT EVERY NUMBER FROM THE CELL TEXT AS WRITTEN and resolve each one.
import re
SPREADS = {r[0]: (r[1], r[2]) for r in
           con.execute("SELECT number_id, spread_min, spread_max FROM dim_headline_number")}

WORDS = {"zero":0,"one":1,"two":2,"three":3,"four":4,"five":5,"six":6,"seven":7,"eight":8,
         "nine":9,"ten":10,"eleven":11,"twelve":12,"thirteen":13,"fourteen":14,"fifteen":15,
         "sixteen":16,"seventeen":17,"eighteen":18,"nineteen":19,"twenty":20,"thirty":30,
         "sixty":60,"hundred":100,"thirty-five":35,"thirty-three":33,"sixty-four":64,"half":0.5,"third":1/3,
         "twice":2,"second":2,"first":1}
# Numbers that are NAMES, not quantities. Listed and reported, never silently dropped.
NAMES = re.compile(r"Bala\s+\d+[a-z]*(?:/\d+)?|Dinan\s+\d+|veg_p05_spatial|veg_p05_mean|"
                   r"\bp05\b|I-\d+|T3-I\d+|\bT\d+\b|\bM\d+b?\b|\bF\d+\b|\bQ\d\b|EPSG:\d+|"
                   r"L-\d+|\bR\d+\b|_bala\d+ca|\b25[- ]met(?:re|er)\b|\b25 m\b")
DATES = re.compile(r"\b\d{1,2} (?:January|February|March|April|May|June|July|August|September|"
                   r"October|November|December) \d{4}\b")
TOKEN = re.compile(r"[-+]?\d[\d,]*(?:\.\d+)?%?(?:st|nd|rd|th)?")
COMPOUND = ["thirty-five", "thirty-three", "sixty-four"]

def numbers_in(text):
    """Every quantity written in the text. Names and dates are removed and reported separately,
    never silently dropped - a number that vanishes from the scan is exactly what RT-1 caught."""
    names = NAMES.findall(text) + DATES.findall(text)
    stripped = DATES.sub(" ", NAMES.sub(" ", text))
    out = []
    for m in TOKEN.finditer(stripped):
        raw = m.group(0)
        v = float(re.sub(r"[,%]|st$|nd$|rd$|th$", "", raw))
        dp = len(raw.split(".")[1].rstrip("%stndrdh")) if "." in raw else 0
        out.append((raw, v, dp, False))
    for w in COMPOUND:                                  # longest first, then blanked, so
        if re.search(rf"\b{w}\b", stripped, re.I):      # "Thirty-five" is not also read as "thirty"
            out.append((w, float(WORDS[w]), 0, True))
            stripped = re.sub(rf"\b{w}\b", " ", stripped, flags=re.I)
    for w, v in WORDS.items():
        if w in COMPOUND: continue
        if re.search(rf"\b{w}\b", stripped, re.I): out.append((w, float(v), 0, True))
    return out, names

# Non-registry quantities, split the way CLAUDE.md already splits result numbers. A PARAMETER is
# the spec and may be written anywhere without citation. A RESULT may not: a cell that carries a
# provenance COLUMN must cite the results it quotes. That split is what lets the uncited-number
# test bite instead of drowning in two dozen legitimate parameters.
PARAMS = {
 35.0:"parameter - 35 water years, 1988-2022 (record length)",
 1988.0:"parameter - first water year of the record",
 2019.0:"parameter - conservation management start (dim_management_zone)",
 2013.0:"illustrative date in a hypothetical, not a measurement",
 64.0:"parameter - 64 management zones (dim_management_zone, 64 rows)",
 115.0:"parameter - 115 paddock-parts (fact_zone_community_part_classification)",
 66.0:"parameter - 66 monitoring plots (dim_plot)",
 4.0:"parameter - 4 conserved paddocks; also the 4 water years since management changed",
 5.0:"parameter - dim_management_zone reserves 5 cropping-history columns",
 3.0:"parameter - 3 grazing regimes / 3 non-treed communities",
 0.50:"parameter - the T13 sweep cut range 0.50 to 1.50 (T13 spec section 5)",
 1.50:"parameter - the T13 sweep cut range 0.50 to 1.50 (T13 spec section 5)",
 1.0:"parameter - 1 pp of threshold, the elasticity numerator (T3 Gate B1)",
 11.0:"parameter - 11 census strata (census_by_zone_stratum)",
 9.0:"parameter - 9 non-treed strata (treed_context_flag=0 AND regime_band<>'context')",
 1080157.0:"parameter - census pixel count (census_by_zone_stratum, verified P3 S2)",
}
RESULTS = {
 988831.0:"census_by_zone_stratum - non-treed scope, verified independently P3 S2",
 91.55:"census_by_zone_stratum - non-treed share, verified independently P3 S2",
 8.0:"census_by_zone_stratum - Floodplain Woodland/Forest share %, verified P3 S2",
 0.46:"census_by_zone_stratum - Other/minor units share %, verified P3 S2",
 2.0:"T13 Gate C - 2 of those 3 in Bala 29ca, verified this session",
 12.0:"T13 Gate C - 12 of the 16 declining parts in the Bala group, verified this session",
 0.5:"derived and stated - r=0.71 gives r-squared 0.504, 'about half'",
 32.1:"t10_bala29ca_aeolian_level_deficit", 24.9:"t10_bala29ca_riverine_level_deficit",
 5.8:"t10_bala29ca_inland_level_deficit",
 32.0:"t10_bala29ca_aeolian_level_deficit (32.1, written to 0 dp)",
 25.0:"t10_bala29ca_riverine_level_deficit (24.9, written to 0 dp)",
 6.0:"t10_bala29ca_inland_level_deficit (5.8, written to 0 dp)",
 15.0:"RPTSCOPE_report_set.csv EXCLUDED rows - 15 standard-grazing sites (R1b, independent count)",
 12641.0:"T3_gateB1_threshold_sweep.csv - persistent area at 70% cut",
 8300.0:"T3_gateB1_threshold_sweep.csv - persistent area at 75% cut",
 4179.0:"T3_gateB1_threshold_sweep.csv - persistent area at 80% cut",
 70.0:"T3_gateB1_threshold_sweep.csv - operational cut", 75.0:"T3_gateB1_threshold_sweep.csv - cut",
 80.0:"T3_gateB1_threshold_sweep.csv - cut", 40.0:"T3_gateB1_threshold_sweep.csv - sweep floor",
 90.0:"T3_gateB1_threshold_sweep.csv - sweep ceiling",
 17.0:"two-floor divergence, spec section 6 / T2_zone_annual_veg_extraction.md section 94",
 31.0:"derived and stated - 2019 minus 1988 is 31 years, 'three decades'",
 86375.0:"census_by_zone_stratum - Floodplain Woodland/Forest pixels, verified P3 S2",
 4951.0:"census_by_zone_stratum - Other/minor units pixels, verified P3 S2",
}
DECLARED_BY_LOC = {   # location-specific sources, checked BEFORE the global table
 "what_we_dont_know Part 1": {
   4.0:"parameter - 4 conserved paddocks (dim_management_zone, treatment not grazed)",
   60.0:"derived and stated - 64 zones less the 4 conserved",
   14.0:"parameter - the 14-day rotational grazing regime (dim_management_zone.treatment)",
   5.0:"derived and stated - rank 61 of 64 is the driest 5%",
   1/3:"L-01 / T13 - Bala 29ca is roughly one third each of three communities",
   8.0:"census_by_zone_stratum - Floodplain Woodland/Forest share 8.00%, verified P3 S2",
   2.0:"T3-I5 - persistent ground floods roughly twice as often as the property average",
   0.0:"live query: drift - the value-drift count is zero"},
 "Start_here claim 6": {2.0:"T13 Gate C - 2 of the 3 all-cut Recovering parts are in Bala 29ca, "
                        "verified against T13_gateC_classification.csv this session",
                   3.0:"T13 Gate C - 3 parts Recovering at every cut 0.50-1.50, verified this session"},
 "Start_here claim 7": {12.0:"T13 Gate C - 12 of the 16 Declining parts are in the Bala group, "
                             "verified this session"},
 "By_question Q3":    {2.0:"T10 cross-sectional residuals - Bala 29ca at -16.8 is the second largest "
                           "shortfall, behind Bala 15 at -17.62 (verified this session)"},
 "By_question Q6":    {2.0:"T13 Gate C - the two wettest years dropped in the robustness re-run",
                  12.0:"T13 Gate C - 12 of the 16 Declining parts are in the Bala group, verified",
                  3.0:"T13 Gate C - 3 parts Recovering at every cut tested, verified this session"},
 "Start_here claim 1": {0.0:"derived and stated - the spread -7.038 to +4.987 brackets zero"},
 "By_question Q2":    {0.0:"derived and stated - the spread -7.038 to +4.987 brackets zero"},
 "what_we_dont_know Part 2": {8.0:"issues log I-40 - eight instances",
   6.0:"issues log I-37 - six numeral collisions", 2.0:"issues log I-43 / I-42 - twice, and one rebuilt check",
   3.0:"issues log I-37 - three unrelated eighteens", 33.0:"Task M Gate A - 33 stale claim sites",
   7.0:"issues log I-40 - seven derivations written and not wired", 4.0:"P1 - four paths written from memory",
   18.0:"issues log I-37 - the three unrelated eighteens", 1.0:"one figure that did not reproduce"},
 # How_we_know's own sentence embeds the live values; they resolve to the queries that produced them
 "How_we_know": {float(LIVE["pinned"]):"live query: pinned", float(LIVE["reproduce"]):"live query: reproduce",
   float(LIVE["drift"]):"live query: drift", float(LIVE["nopath"]):"live query: nopath",
   round(LIVE["coverage"],1):"live query: coverage", 70.0:"T3 Gate B1 pre-registered threshold"},
}
# The T3 page's provenance: the ids the page's Part 1 quotes. Given so a rank resolves to its rank.
T3_IDS = ["ref_paddock_flood_rank_bala26ca","ref_paddock_flood_rank_bala27ca",
          "ref_paddock_flood_rank_bala28ca","ref_paddock_flood_rank_bala29ca",
          "cropping_history_null_count","t10_bala29ca_aeolian_level_deficit",
          "t10_bala29ca_riverine_level_deficit","t10_bala29ca_inland_level_deficit",
          "t13_parts_recovering_count","t13_recovering_survive_drop2wettest","floor_flood_r_64pdk"]

def resolve(loc, v, dp, is_word, ids, contract):
    """Collect EVERY source that could produce this number. One hit is an attribution; more than
    one is coverage without attribution, and the CSV says so rather than picking the first."""
    def eq(cd):
        cd = abs(float(cd))
        return abs(cd - abs(v)) < 1e-9 if is_word else round(cd, dp) == round(abs(v), dp)
    hits, from_id, from_loc, from_param = [], False, False, False
    for nid in ids:
        for cd in [PINS.get(nid)] + list(SPREADS.get(nid, (None, None))):
            if cd is not None and eq(cd): hits.append(nid); from_id = True; break
    for k, src in DECLARED_BY_LOC.get(loc, {}).items():
        if eq(k): hits.append(src); from_loc = True
    for k, src in PARAMS.items():
        if eq(k): hits.append(src); from_param = True
    for k, src in RESULTS.items():
        if eq(k): hits.append(src)
    hits = list(dict.fromkeys(hits))
    if not hits: return "", "UNRESOLVED", 0
    # A cell that cites ids, quoting a number that matches NONE of them and is rescued only by the
    # project-wide table, is RT-1's defect exactly: the number is written but not cited. The global
    # table must not be allowed to absorb it silently - that is how "17" passed before RT-1.
    if contract and not from_id and not from_loc and not from_param:
        return " | ".join(hits[:3]), "DECLARED_UNCITED", len(hits)
    return " | ".join(hits[:3]), ("PINNED" if hits[0] in PINS else "DECLARED"), len(hits)

checks = []
def scan(loc, text, ids, contract=True):
    nums, names = numbers_in(text)
    for raw, v, dp, is_word in nums:
        res, state, n = resolve(loc, v, dp, is_word, ids, contract)
        checks.append(dict(location=loc, number_as_written=raw, resolves_to=res, state=state,
                           n_matching_sources=n, agrees=int(state != "UNRESOLVED")))
    # sorted, NOT set order: Python randomises string hashing per process, so an unsorted set made
    # this file differ between runs with identical inputs - spurious drift in a provenance artefact
    for nm in sorted(set(names)):
        checks.append(dict(location=loc, number_as_written=nm,
                           resolves_to="identifier, name or date - not a quantity",
                           state="NAME_NOT_QUANTITY", n_matching_sources=0, agrees=1))

for n, t, nid in CLAIMS:                     scan(f"Start_here claim {n}", t, nid.split(","))
for k, (qq, a, why, nid) in ANSWERS.items(): scan(f"By_question {k}", why,
                                                  [] if nid.startswith("(none") else nid.split(","))
for a, b in CAUTIONS:                        scan("Two_cautions", b, [], contract=False)
scan("How_we_know", cov + " " + FALSIF, [], contract=False)
for key in ("registered","pinned","reproduce","drift","nopath","figures","tables","reports"):
    checks.append(dict(location="How_we_know", number_as_written=str(LIVE[key]),
                       resolves_to=f"live query: {key}", state="LIVE_QUERY",
                       n_matching_sources=1, agrees=1))
t3 = (PACK/"Gayini_what_we_dont_know.md").read_text(encoding="utf-8")
part1, part2 = t3.split("## Part 2", 1)
# the page carries no provenance COLUMN - its sources are inline prose - so T3_IDS are extra
# candidates, not a contract, and the uncited-result test does not apply to it
scan("what_we_dont_know Part 1", part1, T3_IDS, contract=False)
scan("what_we_dont_know Part 2", part2, [], contract=False)

with open(PACK/"PACK1_final_number_check.csv","w",newline="",encoding="utf-8") as f:
    w=csv.DictWriter(f,fieldnames=list(checks[0].keys())); w.writeheader(); w.writerows(checks)
unres=[x for x in checks if x["state"] in ("UNRESOLVED","DECLARED_UNCITED")]
import collections as _co
print(f"  P4-7 (RT-1 rule - every number AS WRITTEN): {len(checks)} numbers checked, "
      f"{len(unres)} UNRESOLVED")
print("     " + " · ".join(f"{k}={v}" for k,v in _co.Counter(x['state'] for x in checks).items()))
for x in unres: print(f"     *** {x['state']:18s} {x['location']:28s} | "
                      f"{x['number_as_written']:10s} -> {x['resolves_to'][:60]}")
con.close()

# ---------- Ruling N: register the workbook ---------------------------------------------------
if unres:
    raise SystemExit("STOP - unresolved numbers; workbook NOT registered.")
cw = sqlite3.connect(DB); cur = cw.cursor()
try:
    cw.execute("BEGIN")
    cur.execute("INSERT OR REPLACE INTO report_asset(report_asset_id,path,title,report_type,"
                "checksum_sha256,path_exists,qa_status,run_id) VALUES (?,?,?,?,?,?,?,?)",
                ("report_gayini_adrian_pack_xlsx","Output/pack/Gayini_Adrian_pack.xlsx",
                 "Gayini Adrian pack workbook - 5 sheets, content regenerated from live queries",
                 "client_pack", sha50(out), 1, "REVIEW", RUN))
    cur.execute("INSERT OR REPLACE INTO table_asset(table_asset_id,path,title,product,n_rows,"
                "checksum_sha256,path_exists,qa_status,run_id,superseded_flag,framing_label,"
                "provenance_note,support_level) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                ("table_pack1_item_list","Output/pack/PACK1_item_list.csv",
                 "PACK-1 item list - 17 items, 17 files","pack_item_list",len(items),
                 sha50(PACK/"PACK1_item_list.csv"),1,"REVIEW",RUN,0,"census_8058",
                 "Single source for both 00_START_HERE.md and the workbook Contents sheet.","mixed"))
    cw.commit(); print("  COMMIT - workbook + item list registered")
except Exception as e:
    cw.rollback(); cw.close(); raise SystemExit(f"ROLLED BACK: {e}")
cw.close()
post = probe("after")
print(f"\n  report_asset {pre['report_asset']} -> {post['report_asset']}   "
      f"table_asset {pre['table_asset']} -> {post['table_asset']}")
